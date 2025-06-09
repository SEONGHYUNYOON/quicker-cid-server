from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, flash
from datetime import datetime, timedelta, time, date
from models import db, Member, CID, Admin, LoginLog, ApiKey, ApiLog, Backup, BackupSchedule, DailyStats, MemberActivity
import os
from openpyxl import Workbook
from dateutil.parser import parse
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from functools import wraps
from openpyxl.styles import Alignment, Font
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import json
import sqlite3
import shutil
# from apscheduler.schedulers.background import BackgroundScheduler
# from apscheduler.triggers.cron import CronTrigger
from sqlalchemy.sql import func

app = Flask(__name__)

# 환경별 설정 로드
from config import config
import os
config_name = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config[config_name])

# Flask-Login 설정
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'index'

# Rate Limiter 설정
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"]
)

db.init_app(app)

# 로그인 시도 제한 설정
MAX_LOGIN_ATTEMPTS = 5
LOCK_TIME_MINUTES = 30

# 백업 디렉토리 설정
BACKUP_DIR = 'backups'
if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)

# 백업 스케줄러 설정 (배포용으로 비활성화)
# scheduler = BackgroundScheduler()
# scheduler.start()

@login_manager.user_loader
def load_user(user_id):
    return Admin.query.get(int(user_id))

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def log_login_attempt(username, success, ip_address, user_agent):
    log = LoginLog(
        username=username,
        success=success,
        ip_address=ip_address,
        user_agent=user_agent
    )
    db.session.add(log)
    db.session.commit()

def require_api_key(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        api_key = request.headers.get('X-API-Key')
        if not api_key:
            return jsonify({'error': 'API 키가 필요합니다.'}), 401
        
        key = ApiKey.query.filter_by(key=api_key, is_active=True).first()
        if not key:
            return jsonify({'error': '유효하지 않은 API 키입니다.'}), 401
        
        # API 키 사용 시간 업데이트
        key.last_used_at = datetime.utcnow()
        db.session.commit()
        
        # API 요청 로깅
        log = ApiLog(
            api_key_id=key.id,
            endpoint=request.path,
            method=request.method,
            request_data=request.get_data(as_text=True),
            ip_address=request.remote_addr,
            user_agent=request.user_agent.string
        )
        db.session.add(log)
        db.session.commit()
        
        return f(*args, **kwargs)
    return decorated_function

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    password = request.form.get('password')
    admin = Admin.query.filter_by(username='admin').first()
    
    # 관리자 계정이 없으면 생성 (최초 실행 시)
    if not admin:
        admin = Admin(username='admin')
        admin.set_password('4568')  # 초기 비밀번호
        db.session.add(admin)
        db.session.commit()
    
    # 계정 잠금 확인
    if admin.is_locked:
        lock_time = admin.last_attempt + timedelta(minutes=LOCK_TIME_MINUTES)
        if datetime.utcnow() < lock_time:
            remaining_minutes = int((lock_time - datetime.utcnow()).total_seconds() / 60)
            flash(f'계정이 잠겼습니다. {remaining_minutes}분 후에 다시 시도해주세요.')
            return redirect(url_for('index'))
        else:
            admin.is_locked = False
            admin.login_attempts = 0
    
    # 비밀번호 확인
    if admin.check_password(password):
        login_user(admin)
        admin.last_login = datetime.utcnow()
        admin.login_attempts = 0
        db.session.commit()
        
        # 로그인 성공 기록
        log_login_attempt('admin', True, request.remote_addr, request.user_agent.string)
        
        return redirect(url_for('dashboard'))
    else:
        # 로그인 실패 처리
        admin.login_attempts += 1
        admin.last_attempt = datetime.utcnow()
        
        if admin.login_attempts >= MAX_LOGIN_ATTEMPTS:
            admin.is_locked = True
            flash(f'로그인 시도 횟수를 초과했습니다. {LOCK_TIME_MINUTES}분 후에 다시 시도해주세요.')
        else:
            remaining_attempts = MAX_LOGIN_ATTEMPTS - admin.login_attempts
            flash(f'비밀번호가 올바르지 않습니다. 남은 시도 횟수: {remaining_attempts}회')
        
        db.session.commit()
        
        # 로그인 실패 기록
        log_login_attempt('admin', False, request.remote_addr, request.user_agent.string)
        
        return redirect(url_for('index'))

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

@app.route('/change-password', methods=['POST'])
@login_required
def change_password():
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    if not current_user.check_password(current_password):
        return jsonify({'error': '현재 비밀번호가 올바르지 않습니다.'}), 400
    
    if new_password != confirm_password:
        return jsonify({'error': '새 비밀번호가 일치하지 않습니다.'}), 400
    
    if len(new_password) < 4:
        return jsonify({'error': '비밀번호는 최소 4자 이상이어야 합니다.'}), 400
    
    current_user.set_password(new_password)
    db.session.commit()
    
    return jsonify({'message': '비밀번호가 변경되었습니다.'})

@app.route('/dashboard')
@login_required
def dashboard():
    members = Member.query.all()
    return render_template('dashboard.html', members=members)

@app.route('/api/members', methods=['GET'])
@login_required
def get_members():
    members = Member.query.all()
    return jsonify([member.to_dict() for member in members])

@app.route('/api/members/<int:id>', methods=['GET'])
@login_required
def get_member(id):
    member = Member.query.get_or_404(id)
    return jsonify(member.to_dict())

@app.route('/api/members', methods=['POST'])
@login_required
def add_member():
    try:
        data = request.json
        print(f"[DEBUG] Received member data: {data}")
        
        # 필수 필드 검증
        required_fields = ['name', 'phone', 'registration_date', 'expiry_date']
        for field in required_fields:
            if not data or not data.get(field):
                error_msg = f'{field}은(는) 필수 입력 항목입니다.'
                print(f"[DEBUG] Validation error: {error_msg}")
                return jsonify({'error': error_msg}), 400
        
        # 전화번호 중복 검사
        existing_member = Member.query.filter_by(phone=data['phone']).first()
        if existing_member:
            error_msg = '이미 등록된 전화번호입니다.'
            print(f"[DEBUG] Duplicate phone error: {data['phone']}")
            return jsonify({'error': error_msg}), 400
        
        # 날짜 파싱
        try:
            registration_date = parse(data['registration_date'])
            expiry_date = parse(data['expiry_date'])
        except Exception as e:
            error_msg = '날짜 형식이 올바르지 않습니다.'
            print(f"[DEBUG] Date parsing error: {e}")
            return jsonify({'error': error_msg}), 400
        
        member = Member(
            name=data['name'],
            phone=data['phone'],
            registration_date=registration_date,
            expiry_date=expiry_date,
            deposit_amount=data.get('deposit_amount', 0),
            referrer=data.get('referrer', '')
        )
        db.session.add(member)
        db.session.flush()  # member.id를 얻기 위해 flush
        
        for cid_value in data.get('cids', []):
            cid = CID(
                cid_value=cid_value, 
                member_id=member.id,
                is_active=True
            )
            db.session.add(cid)
        
        db.session.commit()
        
        # 활동 기록
        log_member_activity(
            member.id,
            'registration',
            amount=member.deposit_amount,
            details=f'CID 수: {len(data.get("cids", []))}'
        )
        
        print(f"[DEBUG] Member created successfully: ID={member.id}, Name={member.name}")
        return jsonify(member.to_dict())
        
    except Exception as e:
        db.session.rollback()
        error_msg = f'회원 등록 중 오류가 발생했습니다: {str(e)}'
        print(f"[DEBUG] Unexpected error: {e}")
        return jsonify({'error': error_msg}), 500

@app.route('/api/members/<int:id>', methods=['PUT'])
@login_required
def update_member(id):
    member = Member.query.get_or_404(id)
    data = request.json
    
    # 입금액 변경 확인
    if 'deposit_amount' in data and data['deposit_amount'] != member.deposit_amount:
        log_member_activity(
            member.id,
            'deposit',
            amount=data['deposit_amount'] - member.deposit_amount,
            details='입금액 변경'
        )
    
    # 만료일 변경 확인
    if 'expiry_date' in data:
        new_expiry = parse(data['expiry_date'])
        if new_expiry != member.expiry_date:
            log_member_activity(
                member.id,
                'renewal',
                details=f'만료일 변경: {new_expiry.strftime("%Y-%m-%d")}'
            )
    
    member.name = data['name']
    member.phone = data['phone']
    member.registration_date = parse(data['registration_date'])
    member.expiry_date = parse(data['expiry_date'])
    member.deposit_amount = data.get('deposit_amount', 0)
    member.referrer = data.get('referrer', '')
    
    CID.query.filter_by(member_id=member.id).delete()
    
    for cid_value in data.get('cids', []):
        cid = CID(
            cid_value=cid_value, 
            member_id=member.id,
            is_active=True
        )
        db.session.add(cid)
    
    db.session.commit()
    return jsonify(member.to_dict())

@app.route('/api/members/<int:id>', methods=['DELETE'])
@login_required
def delete_member(id):
    member = Member.query.get_or_404(id)
    db.session.delete(member)
    db.session.commit()
    return '', 204

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    cid = data.get('cid')
    pw = data.get('pw')
    
    if not cid or not pw:
        return jsonify({'error': 'Missing credentials'}), 400
    
    cid_record = CID.query.filter_by(cid_value=cid, is_active=True).first()
    if not cid_record:
        return jsonify({'error': 'Invalid CID'}), 401
    
    member = cid_record.member
    if member.expiry_date < datetime.utcnow():
        return jsonify({'error': 'Subscription expired'}), 401
    
    return jsonify({
        'token': 'dummy-token',  # 실제 구현에서는 적절한 토큰 생성 필요
        'expiry_date': member.expiry_date.strftime('%Y-%m-%d')
    })

@app.route('/api/export', methods=['GET'])
@login_required
def export_excel():
    members = Member.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "회원 목록"
    
    # 스타일 설정
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # 헤더 추가
    headers = ['번호', '이름', '전화번호', '등록일', '만료일', '입금액', '추천인', 'CID 목록']
    ws.append(headers)
    
    # 헤더 스타일 적용
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
    
    # 데이터 추가
    for idx, member in enumerate(members, 1):
        # 금액에 천단위 콤마 추가
        deposit_amount = format(member.deposit_amount, ',') + '원' if member.deposit_amount else '0원'
        
        row = [
            idx,
            member.name,
            member.phone,
            member.registration_date.strftime('%Y-%m-%d'),
            member.expiry_date.strftime('%Y-%m-%d'),
            deposit_amount,
            member.referrer,
            ', '.join([cid.cid_value for cid in member.cids])
        ]
        ws.append(row)
        
        # 행 데이터 중앙 정렬
        for cell in ws[idx + 1]:
            cell.alignment = center_alignment
    
    # 열 너비 자동 조정 (최소 15, CID 열은 더 넓게)
    for idx, column in enumerate(ws.columns):
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except:
                pass
        # CID 열은 더 넓게 설정
        if idx == 7:  # CID 열
            adjusted_width = max(max_length + 2, 50)  # 최소 50
        else:
            adjusted_width = max(max_length + 2, 15)  # 최소 15
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # 파일 저장
    filename = f'members_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(filename)
    
    return send_file(filename, as_attachment=True)

# API 키 관리 엔드포인트
@app.route('/api/keys', methods=['GET'])
@login_required
def list_api_keys():
    keys = ApiKey.query.all()
    return jsonify([key.to_dict() for key in keys])

@app.route('/api/keys', methods=['POST'])
@login_required
def create_api_key():
    data = request.json
    if not data or 'name' not in data:
        return jsonify({'error': '이름이 필요합니다.'}), 400
    
    key = ApiKey(
        key=ApiKey.generate_key(),
        name=data['name']
    )
    db.session.add(key)
    db.session.commit()
    
    return jsonify(key.to_dict()), 201

@app.route('/api/keys/<int:id>', methods=['DELETE'])
@login_required
def delete_api_key(id):
    key = ApiKey.query.get_or_404(id)
    key.is_active = False
    db.session.commit()
    return '', 204

# 폰번호 로그인 API 엔드포인트 (메인)
@app.route('/api/v1/login', methods=['POST'])
@require_api_key
def login_by_phone():
    data = request.json
    if not data or 'phone' not in data:
        return jsonify({'error': '전화번호가 필요합니다.'}), 400
    
    phone = data['phone']
    
    # 폰번호로 회원 찾기
    member = Member.query.filter_by(phone=phone).first()
    if not member:
        return jsonify({
            'success': False,
            'error': '등록되지 않은 전화번호입니다.'
        }), 401
    
    # 만료일 확인
    if member.expiry_date < datetime.now().date():
        return jsonify({
            'success': False,
            'error': '이용기간이 만료되었습니다.'
        }), 401
    
    # 해당 회원의 활성 CID 목록 가져오기
    active_cids = [cid.cid_value for cid in member.cids if cid.is_active]
    
    return jsonify({
        'success': True,
        'member': {
            'name': member.name,
            'phone': member.phone,
            'expiry_date': member.expiry_date.strftime('%Y-%m-%d'),
            'cids': active_cids,  # 사용 가능한 CID 목록 (계정 확장용)
            'cid_count': len(active_cids)
        }
    })

# CID 인증 API 엔드포인트 (기존 - 하위 호환성)
@app.route('/api/test/status', methods=['GET'])
def test_status():
    """서버 상태 테스트용 엔드포인트 (인증 불필요)"""
    try:
        import os
        database_url = os.environ.get('DATABASE_URL', 'Not set')
        current_db_uri = app.config.get('SQLALCHEMY_DATABASE_URI', 'Not configured')
        
        member_count = Member.query.count()
        members = Member.query.all()
        member_list = [{'id': m.id, 'name': m.name, 'phone': m.phone} for m in members]
        
        return jsonify({
            'status': 'ok',
            'database': 'connected',
            'database_type': 'PostgreSQL' if 'postgresql://' in current_db_uri else 'SQLite',
            'database_url_set': 'Yes' if database_url != 'Not set' else 'No',
            'current_db_uri': current_db_uri[:50] + '...' if len(current_db_uri) > 50 else current_db_uri,
            'member_count': member_count,
            'members': member_list,
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'database': 'disconnected',
            'error': str(e),
            'timestamp': datetime.utcnow().isoformat()
        }), 500

@app.route('/api/v1/verify', methods=['POST'])
@require_api_key
def verify_cid():
    data = request.json
    if not data or 'cid' not in data:
        return jsonify({'error': 'CID가 필요합니다.'}), 400
    
    cid = CID.query.filter_by(cid_value=data['cid']).first()
    if not cid:
        return jsonify({
            'valid': False,
            'message': '등록되지 않은 CID입니다.'
        })
    
    if not cid.is_active:
        return jsonify({
            'valid': False,
            'message': '비활성화된 CID입니다.'
        })
    
    member = Member.query.get(cid.member_id)
    if member.expiry_date < datetime.now():
        return jsonify({
            'valid': False,
            'message': '사용 기간이 만료되었습니다.'
        })
    
    return jsonify({
        'valid': True,
        'message': '인증되었습니다.',
        'expiry_date': member.expiry_date.strftime('%Y-%m-%d'),
        'member_name': member.name,
        'member_phone': member.phone
    })

@app.route('/api/logs', methods=['GET'])
@login_required
def get_api_logs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    logs = ApiLog.query.order_by(ApiLog.timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'logs': [log.to_dict() for log in logs.items],
        'total': logs.total,
        'pages': logs.pages,
        'current_page': logs.page
    })

def create_backup(description=None, is_auto=False):
    """데이터베이스 백업 생성"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_filename = f'backup_{timestamp}.db'
    backup_path = os.path.join(BACKUP_DIR, backup_filename)
    
    # 데이터베이스 파일 복사
    shutil.copy2('quicker.db', backup_path)
    
    # 백업 정보 저장
    backup = Backup(
        filename=backup_filename,
        size=os.path.getsize(backup_path),
        description=description,
        is_auto=is_auto
    )
    db.session.add(backup)
    db.session.commit()
    
    return backup

def restore_backup(backup_id):
    """백업에서 데이터베이스 복원"""
    backup = Backup.query.get_or_404(backup_id)
    backup_path = os.path.join(BACKUP_DIR, backup.filename)
    
    if not os.path.exists(backup_path):
        raise FileNotFoundError('백업 파일을 찾을 수 없습니다.')
    
    # 데이터베이스 연결 해제
    db.session.remove()
    
    # 현재 데이터베이스 파일 백업
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    current_backup = f'pre_restore_{timestamp}.db'
    shutil.copy2('quicker.db', os.path.join(BACKUP_DIR, current_backup))
    
    # 백업 파일 복원
    shutil.copy2(backup_path, 'quicker.db')
    
    return True

def cleanup_old_backups():
    """오래된 자동 백업 삭제"""
    for schedule in BackupSchedule.query.filter_by(is_active=True).all():
        retention_date = datetime.utcnow() - timedelta(days=schedule.retention_days)
        old_backups = Backup.query.filter(
            Backup.is_auto == True,
            Backup.created_at < retention_date
        ).all()
        
        for backup in old_backups:
            backup_path = os.path.join(BACKUP_DIR, backup.filename)
            if os.path.exists(backup_path):
                os.remove(backup_path)
            db.session.delete(backup)
        
        db.session.commit()

def schedule_backup(schedule_id):
    """백업 스케줄 설정 (배포용으로 비활성화)"""
    # schedule = BackupSchedule.query.get(schedule_id)
    # if not schedule or not schedule.is_active:
    #     return
    # 
    # if schedule.frequency == 'daily':
    #     trigger = CronTrigger(
    #         hour=schedule.time.hour,
    #         minute=schedule.time.minute
    #     )
    # elif schedule.frequency == 'weekly':
    #     trigger = CronTrigger(
    #         day_of_week='mon',
    #         hour=schedule.time.hour,
    #         minute=schedule.time.minute
    #     )
    # elif schedule.frequency == 'monthly':
    #     trigger = CronTrigger(
    #         day=1,
    #         hour=schedule.time.hour,
    #         minute=schedule.time.minute
    #     )
    # else:
    #     return
    # 
    # job_id = f'backup_schedule_{schedule_id}'
    # if job_id in scheduler.get_jobs():
    #     scheduler.remove_job(job_id)
    # 
    # scheduler.add_job(
    #     func=lambda: create_backup(
    #         description=f'자동 백업 ({schedule.frequency})',
    #         is_auto=True
    #     ),
    #     trigger=trigger,
    #     id=job_id,
    #     name=f'Backup Schedule {schedule_id}'
    # )
    pass

# 백업 관련 API 엔드포인트
@app.route('/api/backups', methods=['GET'])
@login_required
def list_backups():
    backups = Backup.query.order_by(Backup.created_at.desc()).all()
    return jsonify([backup.to_dict() for backup in backups])

@app.route('/api/backups', methods=['POST'])
@login_required
def create_backup_api():
    data = request.json or {}
    backup = create_backup(description=data.get('description'))
    return jsonify(backup.to_dict()), 201

@app.route('/api/backups/<int:id>', methods=['DELETE'])
@login_required
def delete_backup(id):
    backup = Backup.query.get_or_404(id)
    
    # 자동 백업은 삭제 불가
    if backup.is_auto:
        return jsonify({'error': '자동 백업은 삭제할 수 없습니다.'}), 400
    
    backup_path = os.path.join(BACKUP_DIR, backup.filename)
    if os.path.exists(backup_path):
        os.remove(backup_path)
    
    db.session.delete(backup)
    db.session.commit()
    
    return '', 204

@app.route('/api/backups/<int:id>/restore', methods=['POST'])
@login_required
def restore_backup_api(id):
    try:
        restore_backup(id)
        return jsonify({'message': '백업이 성공적으로 복원되었습니다.'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/backup-schedules', methods=['GET'])
@login_required
def list_backup_schedules():
    schedules = BackupSchedule.query.all()
    return jsonify([schedule.to_dict() for schedule in schedules])

@app.route('/api/backup-schedules', methods=['POST'])
@login_required
def create_backup_schedule():
    data = request.json
    if not data or 'frequency' not in data or 'time' not in data:
        return jsonify({'error': '필수 정보가 누락되었습니다.'}), 400
    
    try:
        schedule_time = datetime.strptime(data['time'], '%H:%M').time()
    except ValueError:
        return jsonify({'error': '잘못된 시간 형식입니다.'}), 400
    
    schedule = BackupSchedule(
        frequency=data['frequency'],
        time=schedule_time,
        retention_days=data.get('retention_days', 30)
    )
    db.session.add(schedule)
    db.session.commit()
    
    # 백업 스케줄 설정
    schedule_backup(schedule.id)
    
    return jsonify(schedule.to_dict()), 201

@app.route('/api/backup-schedules/<int:id>', methods=['PUT'])
@login_required
def update_backup_schedule(id):
    schedule = BackupSchedule.query.get_or_404(id)
    data = request.json
    
    if 'frequency' in data:
        schedule.frequency = data['frequency']
    if 'time' in data:
        try:
            schedule.time = datetime.strptime(data['time'], '%H:%M').time()
        except ValueError:
            return jsonify({'error': '잘못된 시간 형식입니다.'}), 400
    if 'retention_days' in data:
        schedule.retention_days = data['retention_days']
    if 'is_active' in data:
        schedule.is_active = data['is_active']
    
    db.session.commit()
    
    # 백업 스케줄 업데이트 (배포용으로 비활성화)
    # if schedule.is_active:
    #     schedule_backup(schedule.id)
    # else:
    #     job_id = f'backup_schedule_{schedule.id}'
    #     if job_id in scheduler.get_jobs():
    #         scheduler.remove_job(job_id)
    
    return jsonify(schedule.to_dict())

@app.route('/api/backup-schedules/<int:id>', methods=['DELETE'])
@login_required
def delete_backup_schedule(id):
    schedule = BackupSchedule.query.get_or_404(id)
    
    # 스케줄러에서 작업 제거 (배포용으로 비활성화)
    # job_id = f'backup_schedule_{schedule.id}'
    # if job_id in scheduler.get_jobs():
    #     scheduler.remove_job(job_id)
    
    db.session.delete(schedule)
    db.session.commit()
    
    return '', 204

# 기존 스케줄 복원 (배포용으로 비활성화)
# @app.before_first_request
# def init_backup_schedules():
#     for schedule in BackupSchedule.query.filter_by(is_active=True).all():
#         schedule_backup(schedule.id)
#     
#     # 매일 자정에 오래된 백업 정리
#     scheduler.add_job(
#         func=cleanup_old_backups,
#         trigger=CronTrigger(hour=0, minute=0),
#         id='cleanup_old_backups',
#         name='Cleanup Old Backups'
#     )

# 통계 API 엔드포인트
@app.route('/api/stats/overview', methods=['GET'])
@login_required
def get_stats_overview():
    """전체 통계 개요"""
    days = request.args.get('days', 30, type=int)
    
    return jsonify({
        'members': Member.get_stats(days),
        'cids': CID.get_stats(),
        'api': ApiLog.get_stats(days)
    })

@app.route('/api/stats/daily', methods=['GET'])
@login_required
def get_daily_stats():
    """일일 통계 데이터"""
    days = request.args.get('days', 30, type=int)
    end_date = date.today()
    start_date = end_date - timedelta(days=days)
    
    # 누락된 날짜의 통계 계산
    current_date = start_date
    while current_date <= end_date:
        DailyStats.calculate_stats(current_date)
        current_date += timedelta(days=1)
    
    stats = DailyStats.query.filter(
        DailyStats.date.between(start_date, end_date)
    ).order_by(DailyStats.date.asc()).all()
    
    return jsonify([stat.to_dict() for stat in stats])

@app.route('/api/stats/members/<int:member_id>/activity', methods=['GET'])
@login_required
def get_member_activity(member_id):
    """회원 활동 내역"""
    member = Member.query.get_or_404(member_id)
    activities = MemberActivity.query.filter_by(
        member_id=member_id
    ).order_by(MemberActivity.timestamp.desc()).all()
    
    return jsonify([activity.to_dict() for activity in activities])

@app.route('/api/stats/api-usage', methods=['GET'])
@login_required
def get_api_usage_stats():
    """API 사용 통계"""
    days = request.args.get('days', 30, type=int)
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days)
    
    # API 키별 사용량
    usage_by_key = db.session.query(
        ApiKey.name,
        func.count(ApiLog.id)
    ).join(ApiLog).filter(
        ApiLog.timestamp >= start_date
    ).group_by(ApiKey.id).all()
    
    # 시간대별 사용량
    usage_by_hour = db.session.query(
        func.extract('hour', ApiLog.timestamp),
        func.count(ApiLog.id)
    ).filter(
        ApiLog.timestamp >= start_date
    ).group_by(
        func.extract('hour', ApiLog.timestamp)
    ).all()
    
    # 상태 코드별 분포
    status_distribution = db.session.query(
        ApiLog.status_code,
        func.count(ApiLog.id)
    ).filter(
        ApiLog.timestamp >= start_date
    ).group_by(ApiLog.status_code).all()
    
    return jsonify({
        'by_key': [{'key': k, 'count': c} for k, c in usage_by_key],
        'by_hour': [{'hour': int(h), 'count': c} for h, c in usage_by_hour],
        'status_codes': [{'code': s, 'count': c} for s, c in status_distribution]
    })

# 회원 활동 기록 함수
def log_member_activity(member_id, activity_type, amount=None, details=None):
    """회원 활동 기록"""
    activity = MemberActivity(
        member_id=member_id,
        activity_type=activity_type,
        amount=amount,
        details=details
    )
    db.session.add(activity)
    db.session.commit()

# 일일 통계 계산 스케줄링 (배포용으로 비활성화)
# scheduler.add_job(
#     func=lambda: DailyStats.calculate_stats(date.today()),
#     trigger=CronTrigger(hour=0, minute=5),  # 매일 00:05에 실행
#     id='calculate_daily_stats',
#     name='Calculate Daily Stats'
# )

# 데이터베이스 초기화 (테이블이 없을 때만 - 안전한 방식)
with app.app_context():
    try:
        # Inspector를 사용하여 테이블 존재 확인
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        existing_tables = inspector.get_table_names()
        
        # 필수 테이블들이 없을 때만 생성
        required_tables = ['member', 'cid', 'admin']
        missing_tables = [table for table in required_tables if table not in existing_tables]
        
        if missing_tables:
            print(f"Missing tables detected: {missing_tables}")
            print("Creating database tables...")
            db.create_all()
            print("Database tables created successfully.")
        else:
            print("All required tables exist. Skipping table creation.")
            
    except Exception as e:
        print(f"Database initialization error: {e}")
        # 테이블 확인 실패 시에만 생성 시도
        print("Attempting to create tables...")
        db.create_all()
        print("Database tables created.")

@app.route('/api/test/tables', methods=['GET'])
def test_tables():
    """데이터베이스 테이블 상태 확인"""
    try:
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        existing_tables = inspector.get_table_names()
        
        # Admin 테이블 확인
        admin_exists = 'admin' in existing_tables
        admin_count = Admin.query.count() if admin_exists else 0
        
        return jsonify({
            'status': 'ok',
            'existing_tables': existing_tables,
            'admin_table_exists': admin_exists,
            'admin_count': admin_count,
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e),
            'timestamp': datetime.utcnow().isoformat()
        }), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
